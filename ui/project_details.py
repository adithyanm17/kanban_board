import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from tkcalendar import DateEntry

class ProjectDetailsForm(tk.Frame):
    def __init__(self, parent, db, project):
        super().__init__(parent, bg="white")
        self.db = db
        self.project = project
        self.widgets = {}
        self.setup_ui()

    def setup_ui(self):
        canvas = tk.Canvas(self, bg="white", highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        
        # FIX: Define as self.scrollable_frame so it's accessible throughout the class
        self.scrollable_frame = tk.Frame(canvas, bg="white")

        self.scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def add_field(parent, label_text, attr, row, col, is_date=False):
            tk.Label(parent, text=label_text, bg="white", font=("Arial", 8, "bold")).grid(row=row, column=col*2, sticky="w", padx=(15, 2), pady=8)
            val = getattr(self.project, attr) or ""
            
            if is_date:
                widget = DateEntry(parent, width=20, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
                if val: 
                    try: widget.set_date(val)
                    except: pass
            else:
                widget = tk.Entry(parent, width=25, font=("Arial", 9))
                widget.insert(0, val)
            
            widget.grid(row=row, column=col*2 + 1, sticky="w", padx=(0, 15), pady=8)
            self.widgets[attr] = widget

        def add_text_area(parent, label_text, attr, row, col):
            frame = tk.Frame(parent, bg="white")
            frame.grid(row=row, column=col*4, columnspan=4, padx=15, pady=10, sticky="nsew")
            
            tk.Label(frame, text=label_text, bg="white", font=("Arial", 8, "bold")).pack(anchor="w")
            widget = tk.Text(frame, height=6, width=55, font=("Arial", 9), bd=1, relief="solid")
            widget.insert("1.0", getattr(self.project, attr) or "")
            widget.pack(fill="both", expand=True)
            self.widgets[attr] = widget

        # --- Fields Layout using self.scrollable_frame ---
        add_field(self.scrollable_frame, "Project Name:", "name", 0, 0)
        add_field(self.scrollable_frame, "Customer:", "customer", 0, 1)
        add_field(self.scrollable_frame, "Project Incharge:", "project_manager", 0, 2)
        add_field(self.scrollable_frame, "Total Cost:", "total_cost", 0, 3)

        add_field(self.scrollable_frame, "Part Number:", "part_number", 1, 0)
        add_field(self.scrollable_frame, "Part Name:", "part_name", 1, 1)
        add_field(self.scrollable_frame, "PO Number:", "po_number", 1, 2)
        add_field(self.scrollable_frame, "Work Order:", "wo_number", 1, 3)

        add_field(self.scrollable_frame, "Planned Start:", "start_date", 2, 0, is_date=True)
        add_field(self.scrollable_frame, "Planned End:", "end_date", 2, 1, is_date=True)
        add_field(self.scrollable_frame, "Duration (Hrs):", "estimated_time", 2, 2)
        add_field(self.scrollable_frame, "Due Date:", "due_date", 2, 3, is_date=True)

        add_field(self.scrollable_frame, "PO Date:", "po_date", 3, 0, is_date=True)

        add_text_area(self.scrollable_frame, "Scopes:", "scopes", 4, 0)
        add_text_area(self.scrollable_frame, "Out of Scopes:", "out_of_scopes", 4, 1)

        add_text_area(self.scrollable_frame, "Deliverables:", "deliverables", 5, 0)
        add_text_area(self.scrollable_frame, "Project Description:", "description", 5, 1)

        # --- Action Buttons ---
        button_frame = tk.Frame(self.scrollable_frame, bg="white")
        button_frame.grid(row=6, column=0, columnspan=8, pady=30)

        btn_save = tk.Button(button_frame, text="Save All Project Details", bg="#4CAF50", fg="white", 
                             command=self.save_changes, font=("Arial", 10, "bold"), padx=20, pady=10)
        btn_save.pack(side="left", padx=10)

        btn_export = tk.Button(button_frame, text="Export to Excel", bg="#0078D4", fg="white", 
                               command=self.export_to_excel, font=("Arial", 10, "bold"), padx=20, pady=10)
        btn_export.pack(side="left", padx=10)

    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Project_Report_{self.project.name}.xlsx"
        )
        
        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws1 = wb.active
            ws1.title = "Project Overview"
            
            data_map = [
                ("Project Name", self.project.name),
                ("Customer", self.project.customer),
                ("Project Incharge", self.project.project_manager),
                ("Total Cost", self.project.total_cost),
                ("Planned Start", self.project.start_date),
                ("Planned End", self.project.end_date),
                ("PO Number", self.project.po_number),
                ("Work Order", self.project.wo_number),
                ("Description", self.project.description)
            ]
            
            for i, (label, val) in enumerate(data_map, 1):
                ws1.cell(row=i, column=1, value=label).font = openpyxl.styles.Font(bold=True)
                ws1.cell(row=i, column=2, value=val)

            ws2 = wb.create_sheet("Task List")
            ws2.append(["Title", "Description", "Status", "Assignees"])
            
            all_columns = ["Backlog", "Todo", "WIP", "Testing", "Done", "Approved"]
            for col in all_columns:
                tasks = self.db.get_tasks_by_status(self.project.id, col)
                for t in tasks:
                    assignees = ", ".join(self.db.get_task_assignees(t.id))
                    ws2.append([t.title, t.description, t.status, assignees])

            ws3 = wb.create_sheet("Milestones & Planning")
            ws3.append(["Activity", "Start Date", "End Date"])
            plans = self.db.get_project_plans(self.project.id)
            for p in plans:
                ws3.append([p[2], p[3], p[4]])

            wb.save(file_path)
            messagebox.showinfo("Success", f"Project report exported to:\n{file_path}")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Could not create Excel file: {e}")

    def save_changes(self):
        data = {}
        for key, widget in self.widgets.items():
            if isinstance(widget, tk.Text):
                data[key] = widget.get("1.0", tk.END).strip()
            elif isinstance(widget, DateEntry):
                data[key] = widget.get_date().strftime("%Y-%m-%d")
            else:
                data[key] = widget.get()

        try:
            self.db.update_project(self.project.id, data)
            for key, val in data.items(): 
                setattr(self.project, key, val)
            messagebox.showinfo("Success", "All project details have been updated.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to database: {e}")